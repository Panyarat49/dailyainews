#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
pboat_universe.py — RSS pre-screening funnel for pboat (Panyarat49/dailyainews).

================================  WHAT THIS FILE IS  ================================
The news-GATHERING half of the daily pipeline, written in plain Python. It runs BEFORE
Claude wakes up, casts a wide net over RSS, narrows it to fresh + relevant + trusted
stories, and hands Claude a ready-made candidate list so Claude can spend its time on
judgement and writing instead of hunting around the web.

==============================  WHERE IT FITS (daily)  =============================
  [1] pboat-data.yml  @05:57 BKK ─ runs THIS script ─▶ commits universe_{date}_{stream}.json
  [2] Claude Routine  @07:00 BKK ─ engine Step 0.5 reads that JSON ─▶ writes articles/{date}-{stream}.md
  [3] promote-brief.yml ─ copies the brief from a claude/* branch ─▶ main
  [4] email-notify.yml ─ send_email.py ─▶ Gmail SMTP ─▶ your inbox
  ▸ THIS FILE = STEP 1.

================================  WHAT IT DOES (steps)  ============================
  1. FETCH   ~14 direct RSS feeds + Google News searches (AI/Thai + one per Tier-1 company)
  2. DEDUP   drop the same URL arriving from multiple feeds
  3. KEYWORD keep only AI/tech (ainews) or watchlist-company (watchlist) items; drop noise
  4. CLUSTER group near-identical headlines → cross-outlet corroboration (cluster_size)
  5. TRUSTED drop any publisher not on shared/trusted-sources.md  (default; --all-sources keeps them)
  6. SCORE   multi-factor rank: recency + keyword strength + source role (primary/citation/
             screening) + corroboration + watchlist-company tier; each item keeps a score_breakdown
  7. ENRICH  fetch the FULL article body for the top ~12/stream (browser-UA requests, with a
             Playwright headless-Chromium fallback that also resolves news.google.com links)
  8. WRITE   top ~40 per stream → universe_{YYYY-MM-DD}_{ainews,watchlist}.json
             + a combined ranked workbook → universe_{YYYY-MM-DD}.xlsx and universe-latest.xlsx

Output files (date is dashed to match the engine's {TODAY}):
    .github/scripts/output/universe_{YYYY-MM-DD}_ainews.json      (general AI/tech)
    .github/scripts/output/universe_{YYYY-MM-DD}_watchlist.json   (watchlist companies)
    .github/scripts/output/universe-latest.xlsx                   (ranked workbook, both streams)

The WebFetch-403 bypass: step 6 writes a `body_text` (full article text) into each enriched
candidate. The engine reads that body straight from the committed JSON, so a publisher that
403s Claude's WebFetch tool no longer degrades the brief — the verification can happen against
the funnel-extracted body instead. Items that can't be extracted keep their RSS snippet
(`description`), which the engine still cites at Tier-2.

NOT a trust bypass: the engine re-applies all gates (freshness ≤24h, 7-day dedup,
trusted-source allowlist) before any story appears in a brief. This script replaces Claude's
initial WebSearch discovery pass and pre-fetches the bodies; it does not relax any gate.

Usage:
    python pboat_universe.py                 # both streams, trusted-only, enrich top 12 (default)
    python pboat_universe.py --stream ainews
    python pboat_universe.py --hours 36      # widen the lookback window
    python pboat_universe.py --all-sources   # keep off-allowlist publishers (tagged, not dropped)
    python pboat_universe.py --enrich-n 20   # fetch full bodies for the top 20/stream
    python pboat_universe.py --no-enrich     # skip body extraction (RSS titles+snippets only)
    python pboat_universe.py --no-playwright # requests-only extraction (no Chromium fallback)

Deps (auto-installed on first run): requests feedparser openpyxl
Optional (for the redirect/403 fallback): playwright + chromium  (degrades gracefully if absent)
"""
from __future__ import annotations

import argparse
import html as _html
import importlib
import json
import os
import re
import subprocess
import sys
import time
from datetime import datetime, timezone, timedelta
from pathlib import Path
from urllib.parse import quote


# ── dependency bootstrap (stdlib only above this line) ──────────────────────

def _ensure(pkg: str) -> bool:
    name = pkg.split("[")[0]
    try:
        importlib.import_module(name)
        return True
    except ImportError:
        pass
    print(f"[setup] installing {pkg!r} ...", flush=True)
    r = subprocess.run([sys.executable, "-m", "pip", "install", "--quiet", pkg],
                       capture_output=True, text=True)
    if r.returncode != 0 and "externally-managed-environment" in (r.stderr or ""):
        r = subprocess.run([sys.executable, "-m", "pip", "install", "--quiet",
                            "--break-system-packages", pkg],
                           capture_output=True, text=True)
    importlib.invalidate_caches()
    return r.returncode == 0


for _pkg in ("requests", "feedparser", "openpyxl"):
    if not _ensure(_pkg):
        sys.exit(f"[fatal] could not install {_pkg!r}")

import requests   # noqa: E402
import feedparser # noqa: E402

# Playwright is OPTIONAL (the funnel still runs without it — requests-only extraction +
# RSS-snippet fallback). It is what resolves news.google.com redirect links to the real
# article and recovers full-text bodies from sites that 403 a plain `requests` fetch.
try:
    from playwright.sync_api import sync_playwright  # noqa: E402
    PLAYWRIGHT_OK = True
except Exception:
    sync_playwright = None
    PLAYWRIGHT_OK = False


# ── paths ────────────────────────────────────────────────────────────────────

REPO_ROOT  = Path(__file__).parent.parent.parent   # dailyainews/
SCRIPTS    = Path(__file__).parent                 # .github/scripts/
OUTPUT_DIR = SCRIPTS / "output"
WATCHLIST_PATH    = REPO_ROOT / ".claude/skills/daily-ai-watchlist/reference/watchlist.json"
TRUSTED_SRC_PATH  = REPO_ROOT / ".claude/skills/shared/trusted-sources.md"

UA = {
    "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                   "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"),
    "Accept": "application/rss+xml, application/atom+xml, application/xml, text/xml, */*",
    "Accept-Language": "en-US,en;q=0.9,th;q=0.5",
}

# Same browser identity, but an HTML Accept header for fetching article PAGES (not feeds).
# This is the bypass: a real-browser UA gets a 200 from most publishers that 403 a bot.
ARTICLE_HEADERS = {
    "User-Agent": UA["User-Agent"],
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.8,th;q=0.6",
}

BKK = timezone(timedelta(hours=7))   # Asia/Bangkok UTC+7


# ── AI/tech keyword set for ainews stream ───────────────────────────────────
# Covers: models/labs, chips/hardware, cloud/infra, research, regulation, products.
# Word-boundary aware for Latin terms; substring match for Thai/CJK.

AI_KEYWORDS: list[str] = [
    # labs & models
    "openai", "anthropic", "deepmind", "mistral", "cohere", "xai", "groq", "perplexity",
    "claude", "chatgpt", "gpt", "gemini", "llama", "copilot", "grok",
    # generic AI terms
    "ai", "artificial intelligence", "machine learning", "deep learning",
    "large language model", "llm", "foundation model", "generative ai", "gen ai",
    "multimodal", "agi", "neural network", "transformer", "diffusion model",
    "fine-tuning", "inference", "training run", "benchmark",
    # hardware & infra
    "nvidia", "gpu", "h100", "h200", "blackwell", "gb200",
    "amd", "instinct", "mi300", "mi400",
    "chip", "semiconductor", "wafer", "fab", "tsmc", "hbm", "dram",
    "data center", "cloud ai", "ai infrastructure", "ai accelerator",
    # products & use-cases
    "ai agent", "agentic", "ai assistant", "ai search", "ai coding",
    "robotics", "autonomous", "self-driving", "computer vision",
    # policy & safety
    "ai regulation", "ai safety", "ai governance", "export control", "ai act",
    "ai policy", "responsible ai",
    # investment signals
    "ai funding", "ai investment", "ai acquisition", "ai partnership",
    "ai earnings", "ai revenue",
    # Thai
    "เอไอ", "ปัญญาประดิษฐ์", "โมเดล", "ชิป", "ดาต้าเซ็นเตอร์", "เซมิคอนดักเตอร์",
    # Chinese (for Asia coverage)
    "人工智能", "大模型", "芯片", "英伟达", "谷歌", "OpenAI",
]

# Drop-list: these make a good story look AI-related but aren't (noise reduction).
AI_NOISE: list[str] = [
    "world cup", "fifa", "nba", "nfl", "oscar", "grammy", "celebrity", "concert",
    "accident", "crime", "weather", "horoscope",
]


# ── RSS feed catalog ─────────────────────────────────────────────────────────
# Direct RSS URLs for outlets where feeds are stable.
# These supplement Google News RSS queries (which handle the long tail).

DIRECT_FEEDS: list[dict] = [
    # High-value tech press
    {"name": "TechCrunch",    "url": "https://techcrunch.com/feed/"},
    {"name": "The Verge",     "url": "https://www.theverge.com/rss/index.xml"},
    {"name": "Ars Technica",  "url": "https://feeds.arstechnica.com/arstechnica/technology-lab"},
    {"name": "The Register",  "url": "https://www.theregister.com/headlines.atom"},
    {"name": "Tom's Hardware", "url": "https://www.tomshardware.com/rss.xml"},
    {"name": "VentureBeat",   "url": "https://venturebeat.com/feed/"},
    {"name": "Engadget",      "url": "https://www.engadget.com/rss.xml"},
    {"name": "ZDNet",         "url": "https://www.zdnet.com/news/rss.xml"},
    {"name": "IEEE Spectrum",  "url": "https://spectrum.ieee.org/feeds/feed.rss"},
    # Thai tech
    {"name": "Blognone",      "url": "https://www.blognone.com/node/feed"},
    {"name": "Techsauce",     "url": "https://techsauce.co/feed"},
    {"name": "Brand Inside",  "url": "https://brandinside.asia/feed/"},
    {"name": "The Standard",  "url": "https://thestandard.co/feed/"},
    # Wire / global
    {"name": "CNA Tech",      "url": "https://www.channelnewsasia.com/api/v1/rss-outbound-feed?_format=xml&category=10416"},
]

# Google News RSS queries for broad AI/tech sweeps.
# These catch the top stories across all trusted outlets without needing per-site feeds.
GNEWS_AI_QUERIES: list[str] = [
    "AI artificial intelligence news",
    "LLM large language model",
    "chip semiconductor GPU AI",
    "AI data center cloud computing",
    "AI regulation policy",
]

# Google News RSS queries for Thai-language AI news.
GNEWS_TH_QUERIES: list[str] = [
    "ข่าว AI ปัญญาประดิษฐ์",
    "ชิป เซมิคอนดักเตอร์ เอไอ",
]


def gnews_url(query: str, lang: str = "en", country: str = "US", hours: int = 24) -> str:
    """Build a Google News RSS search URL with a rolling time window."""
    days = max(1, (hours + 23) // 24)
    q = f"{query} when:{days}d"
    return f"https://news.google.com/rss/search?q={quote(q)}&hl={lang}&gl={country}&ceid={country}:{lang}"


# ── helpers ──────────────────────────────────────────────────────────────────

def strip_html(s: str) -> str:
    return re.sub(r"<[^>]+>", " ", _html.unescape(s or "")).strip()


def clean_title(t: str) -> str:
    # Google News appends "  - Publisher Name" — strip it so comparison is clean.
    return re.sub(r"\s{2,}-\s+[^-]+$", "", (t or "").strip()).strip()


def parse_feed(content: bytes | str, hours: int) -> list[dict]:
    """Parse an RSS/Atom feed and return items published within the lookback window."""
    import calendar
    feed = feedparser.parse(content)
    cutoff = time.time() - hours * 3600
    out: list[dict] = []
    for e in feed.entries:
        tp = e.get("published_parsed") or e.get("updated_parsed")
        ts = calendar.timegm(tp) if tp else None
        if ts is not None and ts < cutoff:
            continue  # older than window
        age_h = round((time.time() - ts) / 3600, 1) if ts else None
        body = ""
        if e.get("content"):
            try:
                body = e["content"][0].get("value", "") or ""
            except Exception:
                body = ""
        desc = strip_html(body or e.get("summary") or e.get("description") or "")[:500]
        link = e.get("link") or ""

        # Google News RSS items carry a <source url="https://realpublisher.com">Name</source>
        # element — feedparser exposes it as entry.source. Use it to recover the REAL
        # publisher (the link itself is a news.google.com redirect), with no extra HTTP call.
        src = e.get("source") or {}
        publisher_name   = (src.get("title") or "").strip()
        publisher_href   = (src.get("href") or "").strip()
        publisher_domain = source_domain(publisher_href) if publisher_href else ""

        raw_title = (e.get("title") or "").strip()
        # Google News appends " - {Publisher}" to titles. If we know the publisher, strip
        # exactly that suffix (precise — won't eat legit " - " inside a real headline).
        if publisher_name and raw_title.endswith(f" - {publisher_name}"):
            title = raw_title[: -(len(publisher_name) + 3)].strip()
        else:
            title = clean_title(raw_title)

        if not title or not link:
            continue
        out.append({
            "title":            title,
            "url":              link,
            "description":      desc,
            "published_raw":    e.get("published") or e.get("updated") or "",
            "age_h":            age_h,
            "has_timestamp":    tp is not None,
            "publisher":        publisher_name,
            "publisher_domain": publisher_domain,
        })
    return out


def fetch_feed(url: str, hours: int, timeout: int = 15) -> tuple[list[dict], str]:
    """Fetch one RSS feed URL. Returns (items, error_str)."""
    try:
        r = requests.get(url, headers=UA, timeout=timeout)
        if r.status_code != 200:
            return [], f"HTTP {r.status_code}"
        items = parse_feed(r.content, hours)
        return items, ""
    except Exception as e:
        return [], f"{type(e).__name__}: {e}"


def source_domain(url: str) -> str:
    m = re.search(r"https?://(?:www\.)?([^/]+)", url or "")
    return m.group(1).lower() if m else ""


# ── full-text extraction (the WebFetch-403 bypass) ──────────────────────────
# Mirrors P'Jah's funnel: the Python side fetches the article BODY with a real-browser
# User-Agent (and a Playwright headless-Chromium fallback) and writes it into the universe
# JSON. Claude then reads `body_text` from the committed JSON instead of calling the
# WebFetch tool — so a publisher that 403s Claude's WebFetch no longer degrades the brief.

def is_gnews(url: str) -> bool:
    """A Google News RSS link is a news.google.com redirect, NOT the real article. Plain
    `requests` usually can't follow it (the redirect is JS); Playwright can."""
    return "news.google.com" in (url or "")


def extract_article(url: str, max_chars: int = 1500, timeout: int = 12) -> str:
    """Best-effort full-text grab for ONE article via browser-UA `requests`. Returns plain
    text, or '' if blocked (403)/empty/JS-walled. Never raises."""
    if not url:
        return ""
    try:
        r = requests.get(url, headers=ARTICLE_HEADERS, timeout=timeout, allow_redirects=True)
        if r.status_code != 200 or not r.text:
            return ""
        h = re.sub(r"(?is)<(script|style|noscript|nav|header|footer|aside|form)[^>]*>.*?</\1>",
                   " ", r.text)
        paras = re.findall(r"(?is)<p[^>]*>(.*?)</p>", h)
        text = " ".join(strip_html(p) for p in paras) if paras else strip_html(h)
        text = re.sub(r"\s+", " ", text).strip()
        return text[:max_chars]
    except Exception:
        return ""


_PW_READY = False  # one-time Chromium availability check (lazy — only if we actually use PW)


def ensure_chromium() -> None:
    """Make sure a headless Chromium is installed (one-time ~150MB download). Flips
    PLAYWRIGHT_OK off if it can't be made to work, so callers degrade to requests-only."""
    global PLAYWRIGHT_OK, _PW_READY
    if not PLAYWRIGHT_OK or _PW_READY:
        return
    try:
        with sync_playwright() as p:
            p.chromium.launch(headless=True).close()
        _PW_READY = True
    except Exception:
        print("[setup] downloading Playwright Chromium (one-time ~150MB) ...", flush=True)
        try:
            subprocess.run([sys.executable, "-m", "playwright", "install", "chromium"])
            with sync_playwright() as p:
                p.chromium.launch(headless=True).close()
            _PW_READY = True
        except Exception as e:
            print(f"[setup] Chromium unavailable ({e}); Playwright extraction skipped.", flush=True)
            PLAYWRIGHT_OK = False


def pw_fetch(url: str, max_chars: int = 1500, timeout_ms: int = 30000) -> tuple[str, str, str]:
    """Headless-Chromium fetch. Follows JS redirects (so a news.google.com link lands on the
    REAL article), returns (body_text, resolved_url, published_or_empty). ('', '', '') on
    failure. This is the recovery path for 403s and Google News redirects."""
    if not PLAYWRIGHT_OK:
        return "", "", ""
    try:
        with sync_playwright() as p:
            b = p.chromium.launch(headless=True)
            ctx = b.new_context(user_agent=UA["User-Agent"], locale="en-US")
            pg = ctx.new_page()
            try:
                pg.goto(url, wait_until="domcontentloaded", timeout=timeout_ms)
                # a Google News redirect resolves via JS — give it a beat to land on the article
                try:
                    pg.wait_for_url(lambda u: "news.google.com" not in u, timeout=8000)
                except Exception:
                    pass
                resolved = pg.url or ""
                body = ""
                for sel in ("article", "main", "[role=main]"):
                    el = pg.query_selector(sel)
                    if el:
                        t = el.inner_text()
                        if t and len(t) > 200:
                            body = t
                            break
                if not body:
                    body = pg.inner_text("body")
                published = ""
                for sel, attr in (('meta[property="article:published_time"]', "content"),
                                  ('meta[itemprop="datePublished"]', "content"),
                                  ('meta[name="date"]', "content"),
                                  ("time[datetime]", "datetime")):
                    el = pg.query_selector(sel)
                    if el:
                        v = el.get_attribute(attr)
                        if v and v.strip():
                            published = v.strip()
                            break
                body = re.sub(r"\s+", " ", body or "").strip()[:max_chars]
                return body, resolved, published
            finally:
                b.close()
    except Exception:
        return "", "", ""


def enrich_item(it: dict, use_playwright: bool = True) -> dict:
    """Add full-text `body_text` to ONE candidate. Strategy:
      • real publisher URL → try requests first (cheap); fall back to Playwright on 403/empty.
      • news.google.com redirect → go straight to Playwright (requests can't follow it).
    Sets body_text, extract_status (ok|blocked|skipped), and resolved_url. Never raises."""
    url = it.get("url", "")
    it.setdefault("resolved_url", "")
    if not url:
        it["body_text"], it["extract_status"] = "", "skipped"
        return it

    body = "" if is_gnews(url) else extract_article(url)
    if body:
        it["body_text"], it["extract_status"] = body, "ok"
        return it

    if use_playwright and PLAYWRIGHT_OK:
        ensure_chromium()
        pw_body, resolved, _pub = pw_fetch(url)
        if resolved and "news.google.com" not in resolved:
            it["resolved_url"] = resolved
        if pw_body and len(pw_body) > 200:
            it["body_text"], it["extract_status"] = pw_body, "ok"
            return it

    # nothing better than the RSS description — leave it for the engine's Tier-2 snippet path
    it["body_text"], it["extract_status"] = "", "blocked"
    return it


# ── keyword matching ─────────────────────────────────────────────────────────

_WB = r"[^\W_]"  # Unicode word character (letter or digit, not underscore/punct)


def _kw_hit(kw: str, text: str) -> bool:
    """Word-boundary match for Latin keywords; substring match for Thai/CJK."""
    if re.search(r"[a-z0-9]", kw):
        return bool(re.search(
            r"(?<!" + _WB + r")" + re.escape(kw) + r"(?!" + _WB + r")", text
        ))
    return kw in text


def keyword_hits(text: str, keywords: list[str]) -> list[str]:
    """Return subset of `keywords` that match `text` (lowercased)."""
    low = text.lower()
    return [k for k in keywords if _kw_hit(k, low)]


def is_noise(text: str) -> bool:
    low = text.lower()
    return any(_kw_hit(k, low) for k in AI_NOISE)


# ── clustering (cross-outlet corroboration) ─────────────────────────────────
# Same idea as P'Jah's funnel: group near-identical headlines so a story carried by
# several outlets scores higher (corroboration), and so the Excel can show how many
# outlets ran it. Jaccard overlap of title tokens; Thai/CJK-aware tokenizer.

def tokenize(t: str) -> set[str]:
    # Thai block first (its tone/vowel marks are Unicode "non-word", so a generic \w split
    # would shatter Thai words); then digits; then letter-runs of any script.
    return set(re.findall(r"[฀-๿]+|[0-9]+|[^\W\d_]+", (t or "").lower()))


def cluster(items: list[dict], thresh: float = 0.5) -> list[dict]:
    """Tag each item with `cluster_id` + `cluster_size` (how many items share its story)."""
    clusters: list[list] = []   # each: [token_set, [indices]]
    for i, it in enumerate(items):
        toks, placed = tokenize(it.get("title", "")), False
        for cl in clusters:
            inter, union = len(toks & cl[0]), (len(toks | cl[0]) or 1)
            if inter / union >= thresh:
                cl[0].update(toks); cl[1].append(i); placed = True; break
        if not placed:
            clusters.append([set(toks), [i]])
    for cid, cl in enumerate(clusters):
        for idx in cl[1]:
            items[idx]["cluster_id"]   = cid
            items[idx]["cluster_size"] = len(cl[1])
    return items


# ── scoring ──────────────────────────────────────────────────────────────────
# Multi-factor relevance score (mirrors P'Jah's structure, adapted to AI/tech news):
#   score = recency + signal(keyword hits) + source role + corroboration + watchlist tier
# Each item also carries a `score_breakdown` string so the ranking is auditable in the Excel.
# Claude still re-applies the editorial gates, but it now LEADS selection with this score.

ROLE_W = {"primary": 2.0, "citation": 1.2, "screening": 0.5}   # source authority weight
WL_TIER_W = {1: 1.5, 2: 0.8}                                   # watchlist company-tier boost

SCORING_README = {
    "_README": "score = sum of the components below. Higher = more brief-worthy. "
               "Candidates are pre-sorted by score desc; each item's `score_breakdown` shows its parts.",
    "recency":   "0–1h = 2.3 (breaking) · 1–3h = 2.0 (fresh) · then linear decay to 0 by 24h · unknown = 0.5",
    "signal":    "AI/tech (or watchlist) keyword hits — match_count capped at 5 × 0.6",
    "source":    "outlet authority from trusted-sources.md: primary (company/lab) 2.0 · citation 1.2 · screening 0.5 · other 1.0",
    "corrob":    "cross-outlet corroboration — how many outlets ran the same headline (cluster_size − 1, capped 3) × 0.8",
    "watchlist": "watchlist stream only: matched company tier — Tier-1 +1.5 · Tier-2 +0.8",
}


def score_item(item: dict, stream: str) -> tuple[float, str]:
    """Return (score, breakdown). Reads match_count, age_h, cluster_size, source_role, and
    (watchlist stream) company_tier — all set by earlier phases."""
    # recency bands: 0–1h breaking premium · 1–3h fresh plateau · linear decay to 0 by 24h
    age = item.get("age_h")
    if age is None:
        s_rec = 0.5                              # unknown timestamp — mild, not zero
    elif age <= 1:
        s_rec = 2.3
    elif age <= 3:
        s_rec = 2.0
    else:
        s_rec = round(max(0.0, (24 - age) / 21) * 2, 2)
    s_sig  = round(min(item.get("match_count", 0), 5) * 0.6, 2)         # keyword/signal strength
    s_src  = ROLE_W.get(item.get("source_role", ""), 1.0)              # source authority
    s_corr = round(min(item.get("cluster_size", 1) - 1, 3) * 0.8, 2)   # cross-outlet corroboration
    s_wl   = WL_TIER_W.get(item.get("company_tier", 0), 0.0) if stream == "watchlist" else 0.0
    total  = round(s_rec + s_sig + s_src + s_corr + s_wl, 2)
    bd = f"rec{s_rec}+sig{s_sig}+src{s_src}+corr{s_corr}" + (f"+wl{s_wl}" if s_wl else "")
    return total, bd


# ── deduplication ────────────────────────────────────────────────────────────

def dedup_url(url: str) -> str:
    """Normalise URL for dedup: strip scheme, www, trailing slash, query params."""
    u = re.sub(r"^https?://(?:www\.)?", "", url.lower()).rstrip("/")
    u = re.sub(r"\?.*$", "", u)
    return u


# ── watchlist keyword extraction ─────────────────────────────────────────────

def load_watchlist_keywords(path: Path) -> list[str]:
    """
    Read watchlist.json and return a flat, deduplicated list of all keywords
    and cn_terms across all tiers. Used for watchlist stream keyword matching.
    """
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception as e:
        print(f"[warn] could not read watchlist.json: {e}", flush=True)
        return []
    kws: list[str] = []
    for tier_companies in data.get("tiers", {}).values():
        for co in tier_companies:
            kws.extend(co.get("keywords", []))
            kws.extend(co.get("cn_terms", []))
    # Lowercase everything, drop empties
    return list({k.lower() for k in kws if k.strip()})


def load_watchlist_gnews_queries(path: Path) -> list[str]:
    """
    Build Google News queries per Tier-1 company for targeted watchlist sweeps.
    Returns a list of queries like '"Nvidia" OR "Jensen Huang" AI'.
    """
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return []
    queries: list[str] = []
    for co in data.get("tiers", {}).get("1", []):
        kws = co.get("keywords", [])
        if not kws:
            continue
        # Take the first 2 keywords (company name + main product/person) to keep query tight
        terms = " OR ".join(f'"{k}"' for k in kws[:2])
        queries.append(f"({terms}) AI news")
    return queries


def load_watchlist_companies(path: Path) -> list[dict]:
    """Per-company records for tagging the watchlist stream: which company a story is about
    and its tier (1 > 2). Returns [{company, tier, keywords[], cn_terms[]}] (terms lowercased)."""
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return []
    out: list[dict] = []
    for tier_str, companies in data.get("tiers", {}).items():
        try:
            tier = int(tier_str)
        except (TypeError, ValueError):
            tier = 9
        for co in companies:
            out.append({
                "company":  co.get("company", ""),
                "tier":     tier,
                "keywords": [k.lower() for k in co.get("keywords", []) if k.strip()],
                "cn_terms": [k.lower() for k in co.get("cn_terms", []) if k.strip()],
            })
    return out


def tag_watchlist(item: dict, companies: list[dict]) -> None:
    """Set `matched_company` + `company_tier` (best = lowest tier number) on a watchlist item.
    Used both to boost the score and to build the Excel 'by company' tab."""
    text = (item.get("title", "") + " " + item.get("description", "")).lower()
    best: tuple[int, str] | None = None
    for co in companies:
        if any(_kw_hit(k, text) for k in (co["keywords"] + co["cn_terms"])):
            if best is None or co["tier"] < best[0]:
                best = (co["tier"], co["company"])
    item["matched_company"], item["company_tier"] = (best[1], best[0]) if best else ("", 0)


# ── trusted-source allowlist ─────────────────────────────────────────────────

def load_trusted_domains(path: Path) -> set[str]:
    """
    Extract the set of allow-listed domains from trusted-sources.md. Scans only the
    source bullet lines (`- **Name** — domain …`), grabbing every domain-like token on
    each — including parenthetical extras like `(+ blogs.nvidia.com)`. Paths after the
    domain (e.g. `cnn.com/business`) are ignored; only the host is kept.
    """
    dom_re = re.compile(r"\b((?:[a-z0-9-]+\.)+[a-z]{2,})\b", re.I)
    domains: set[str] = set()
    try:
        lines = path.read_text(encoding="utf-8").splitlines()
    except Exception as e:
        print(f"[warn] could not read trusted-sources.md: {e}", flush=True)
        return domains
    for line in lines:
        if not line.lstrip().startswith("- **"):
            continue
        for m in dom_re.findall(line):
            domains.add(m.lower().strip("."))
    return domains


def is_trusted(domain: str, allowlist: set[str]) -> bool:
    """True if `domain` equals, or is a subdomain of, any allow-listed domain."""
    d = (domain or "").lower()
    if not d:
        return False
    return any(d == a or d.endswith("." + a) for a in allowlist)


def load_source_roles(path: Path) -> dict[str, str]:
    """Map each trusted domain → its role from the trusted-sources.md section it sits under:
    'primary' (A/A2 company + lab pages), 'citation' (B open press/wire), 'screening' (C
    paywalled). Drives the source-authority component of the score (ROLE_W)."""
    dom_re = re.compile(r"\b((?:[a-z0-9-]+\.)+[a-z]{2,})\b", re.I)
    roles: dict[str, str] = {}
    role = ""
    try:
        lines = path.read_text(encoding="utf-8").splitlines()
    except Exception:
        return roles
    for line in lines:
        s = line.strip()
        if s.startswith("## "):
            u = s.upper()
            role = ("primary" if "PRIMARY" in u else
                    "citation" if "CITATION" in u else
                    "screening" if "SCREENING" in u else "")
            continue
        if role and s.startswith("- **"):
            for m in dom_re.findall(line):
                roles.setdefault(m.lower().strip("."), role)   # first (most specific) section wins
    return roles


def source_role(domain: str, roles: dict[str, str]) -> str:
    """Role for a domain (exact or subdomain match); '' if unknown."""
    d = (domain or "").lower()
    if not d:
        return ""
    if d in roles:
        return roles[d]
    for dom, r in roles.items():
        if d.endswith("." + dom):
            return r
    return ""


# ── main pipeline ─────────────────────────────────────────────────────────────

def run_stream(
    stream: str,
    hours: int,
    watchlist_keywords: list[str],
    watchlist_gnews: list[str],
    trusted_domains: set[str],
    source_roles: dict[str, str],
    watchlist_companies: list[dict],
    trusted_only: bool = True,
    top_n: int = 40,
    enrich_n: int = 12,
    use_playwright: bool = True,
) -> dict:
    """
    Fetch, filter, score, and return the candidate pool for one stream.
    stream: "ainews" | "watchlist"
    enrich_n: how many top-scored candidates get a full-text `body_text` (0 = none).
    """
    print(f"\n{'='*60}", flush=True)
    print(f"[{stream}] starting (window={hours}h)", flush=True)

    if stream == "ainews":
        keywords      = AI_KEYWORDS
        extra_queries = GNEWS_AI_QUERIES + GNEWS_TH_QUERIES
        gnews_lang_pairs = [("en", "US")] * len(GNEWS_AI_QUERIES) + [("th", "TH")] * len(GNEWS_TH_QUERIES)
    else:  # watchlist
        keywords      = watchlist_keywords
        extra_queries = watchlist_gnews
        gnews_lang_pairs = [("en", "US")] * len(watchlist_gnews)

    # ── Phase 1: collect raw items from all feeds ──
    raw: list[dict] = []
    ok_count = 0

    # 1a. Direct RSS feeds (same for both streams — they cover all trusted outlets)
    print(f"[{stream}] fetching {len(DIRECT_FEEDS)} direct feeds ...", flush=True)
    for feed in DIRECT_FEEDS:
        items, err = fetch_feed(feed["url"], hours)
        if err:
            print(f"  [skip] {feed['name']}: {err}", flush=True)
        else:
            ok_count += 1
            for it in items:
                it["feed_name"] = feed["name"]
            raw.extend(items)
        time.sleep(0.25)  # be polite

    # 1b. Google News RSS sweeps (topic + watchlist per-company)
    print(f"[{stream}] fetching {len(extra_queries)} Google News queries ...", flush=True)
    for i, query in enumerate(extra_queries):
        lang, country = gnews_lang_pairs[i] if i < len(gnews_lang_pairs) else ("en", "US")
        url = gnews_url(query, lang, country, hours)
        items, err = fetch_feed(url, hours)
        if err:
            print(f"  [skip] GNews '{query[:40]}': {err}", flush=True)
        else:
            ok_count += 1
            for it in items:
                it["feed_name"] = f"GNews:{query[:30]}"
            raw.extend(items)
        time.sleep(0.3)

    total_raw = len(raw)
    print(f"[{stream}] raw items: {total_raw} (from {ok_count} feeds)", flush=True)

    # ── Phase 2: dedup by URL ──
    seen_urls: set[str] = set()
    deduped: list[dict] = []
    for it in raw:
        key = dedup_url(it["url"])
        if key and key not in seen_urls:
            seen_urls.add(key)
            deduped.append(it)

    print(f"[{stream}] after URL dedup: {len(deduped)}", flush=True)

    # ── Phase 3: keyword filter + noise gate ──
    candidates: list[dict] = []
    for it in deduped:
        text = it["title"] + " " + it.get("description", "")
        if is_noise(text):
            continue
        matched = keyword_hits(text, keywords)
        if not matched:
            continue
        it["keywords_matched"] = matched
        it["match_count"]      = len(matched)
        # Prefer the REAL publisher domain (from the feed's <source> element) over the
        # link domain, which for Google News items is just the news.google.com redirect.
        it["source"]           = it.get("publisher_domain") or source_domain(it["url"])
        candidates.append(it)

    n_after_keyword = len(candidates)
    print(f"[{stream}] after keyword filter: {n_after_keyword}", flush=True)

    # ── Phase 3.4: cluster near-identical headlines → cross-outlet corroboration ──
    cluster(candidates)

    # ── Phase 3.45: tag watchlist company + tier (watchlist stream only) ──
    if stream == "watchlist":
        for it in candidates:
            tag_watchlist(it, watchlist_companies)

    # ── Phase 3.5: trusted-source allowlist (trusted-sources.md) ──
    # Tag every candidate; by default DROP off-allowlist publishers so the pool Claude
    # receives only contains citeable outlets. `--all-sources` keeps everything (still
    # tagged) for wide discovery. On thin days the engine's Step 0.5 fallback supplements
    # with WebSearch, so filtering here is safe.
    for it in candidates:
        it["on_allowlist"] = is_trusted(it["source"], trusted_domains)
    trusted_n = sum(1 for it in candidates if it["on_allowlist"])
    if trusted_only:
        before = len(candidates)
        candidates = [it for it in candidates if it["on_allowlist"]]
        print(f"[{stream}] trusted-source filter: kept {len(candidates)} / {before} "
              f"(dropped {before - len(candidates)} off-allowlist)", flush=True)
    else:
        print(f"[{stream}] allowlist annotation only: {trusted_n}/{len(candidates)} on trusted-sources", flush=True)
    n_after_trusted = len(candidates)

    # ── Phase 3.6: score (now that cluster_size, source role, and company tier are known) ──
    for it in candidates:
        it["source_role"] = source_role(it["source"], source_roles)
        it["score"], it["score_breakdown"] = score_item(it, stream)

    # ── Phase 4: sort by score desc, take top_n ──
    candidates.sort(key=lambda x: x["score"], reverse=True)
    candidates = candidates[:top_n]

    # ── Phase 4.5: full-text enrichment of the top picks (the WebFetch-403 bypass) ──
    # Only the most likely finalists are fetched (like P'Jah) — fetching all 40 would be
    # slow and mostly wasted on stories that won't be selected. Each enriched item gets a
    # `body_text` so the engine can verify + summarize WITHOUT a live WebFetch.
    n_enrich = min(enrich_n, len(candidates))
    if n_enrich > 0:
        print(f"[{stream}] enriching top {n_enrich} with full article text "
              f"(playwright={'on' if (use_playwright and PLAYWRIGHT_OK) else 'off'}) ...", flush=True)
        n_ok = 0
        for it in candidates[:n_enrich]:
            enrich_item(it, use_playwright=use_playwright)
            if it.get("extract_status") == "ok":
                n_ok += 1
            time.sleep(0.3)  # be polite
        print(f"[{stream}] enrichment: {n_ok}/{n_enrich} got full body "
              f"({n_enrich - n_ok} fell back to RSS snippet)", flush=True)

    # ── Phase 5: clean up for JSON output ──
    output_items: list[dict] = []
    for it in candidates:
        output_items.append({
            "title":           it["title"],
            "url":             it["url"],
            "resolved_url":    it.get("resolved_url", ""),
            "source":          it.get("source", ""),
            "publisher":       it.get("publisher", ""),
            "on_allowlist":    it.get("on_allowlist", False),
            "feed_name":       it.get("feed_name", ""),
            "published_raw":   it.get("published_raw", ""),
            "age_h":           it.get("age_h"),
            "has_timestamp":   it.get("has_timestamp", False),
            "source_role":     it.get("source_role", ""),
            "description":     it.get("description", "")[:300],
            "body_text":       it.get("body_text", ""),
            "extract_status":  it.get("extract_status", "skipped"),
            "keywords_matched": it.get("keywords_matched", [])[:8],  # trim for readability
            "match_count":     it.get("match_count", 0),
            "cluster_size":    it.get("cluster_size", 1),
            "matched_company": it.get("matched_company", ""),
            "company_tier":    it.get("company_tier", 0),
            "score":           it.get("score", 0.0),
            "score_breakdown": it.get("score_breakdown", ""),
        })

    now_bkk = datetime.now(BKK)
    result = {
        "generated_at":  now_bkk.isoformat(),
        "date":          now_bkk.strftime("%Y-%m-%d"),
        "stream":        stream,
        "window_hours":  hours,
        # Self-documenting score key (kept near the top so it's copy-pasteable). Candidates are
        # pre-sorted by `score` desc — the writer LEADS selection with this, then applies its gates.
        "scoring": SCORING_README,
        "stats": {
            "feeds_ok":            ok_count,
            "items_raw":           total_raw,
            "items_after_dedup":   len(deduped),
            "items_after_keyword": n_after_keyword,
            "items_after_trusted": n_after_trusted,
            "trusted_filter":      trusted_only,
            "items_enriched":      sum(1 for it in output_items if it.get("extract_status") == "ok"),
            "playwright":          bool(use_playwright and PLAYWRIGHT_OK),
        },
        "candidates": output_items,
    }
    print(f"[{stream}] {len(output_items)} candidates written", flush=True)
    return result


# ── Excel export ───────────────────────────────────────────────────────────
# One combined workbook (universe-latest.xlsx) with, per stream: an "all" tab sorted by
# score, a "top" tab of the leaders, and (watchlist) a "by company" tab. Mirrors P'Jah's
# universe Excel: styled header, frozen header row, auto-filter, clickable URLs.

NAVY, WHITE, GREY = "1F3B5C", "FFFFFF", "808080"
LINK_BLUE = "0563C1"
STREAM_LABEL = {"ainews": "AI news", "watchlist": "Watchlist"}


def export_universe_xlsx(results: dict[str, dict], path: Path) -> Path:
    """Write the combined ranked-universe workbook from per-stream result dicts."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter as gcl

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # drop the default empty sheet; we add our own

    def header(ws, cols):
        for c, (label, width) in enumerate(cols, 1):
            cell = ws.cell(row=1, column=c, value=label)
            cell.font = Font(bold=True, color=WHITE, size=10)
            cell.fill = PatternFill("solid", fgColor=NAVY)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[gcl(c)].width = width
        ws.freeze_panes = "A2"

    def put(ws, r, c, value, *, wrap=False, link=False):
        cell = ws.cell(row=r, column=c, value=value)
        cell.alignment = Alignment(vertical="top", wrap_text=wrap)
        cell.font = Font(size=9)
        if link and str(value).startswith("http"):
            cell.hyperlink = value
            cell.font = Font(size=9, color=LINK_BLUE, underline="single")
        return cell

    def best_url(it):
        return it.get("resolved_url") or it.get("url") or ""

    def add_all_tab(stream, cands):
        wl = stream == "watchlist"
        cols = [("Score", 8), ("Corrob", 7), ("Role", 9), ("Source", 22)]
        if wl:
            cols += [("Company", 16), ("Tier", 6)]
        cols += [("Headline", 60), ("Snippet", 50), ("Keywords", 26),
                 ("Age(h)", 7), ("Published", 22), ("Verified", 9), ("URL", 14)]
        ws = wb.create_sheet(f"{STREAM_LABEL[stream]} — all"[:31])
        header(ws, cols)
        for r, it in enumerate(cands, start=2):
            c = 1
            put(ws, r, c, it.get("score", 0)); c += 1
            put(ws, r, c, it.get("cluster_size", 1)); c += 1
            put(ws, r, c, it.get("source_role", "")); c += 1
            put(ws, r, c, it.get("source", "")); c += 1
            if wl:
                put(ws, r, c, it.get("matched_company", "")); c += 1
                put(ws, r, c, it.get("company_tier") or ""); c += 1
            put(ws, r, c, it.get("title", ""), wrap=True); c += 1
            put(ws, r, c, (it.get("body_text") or it.get("description") or "")[:400], wrap=True); c += 1
            put(ws, r, c, ", ".join(it.get("keywords_matched", []))); c += 1
            put(ws, r, c, it.get("age_h")); c += 1
            put(ws, r, c, it.get("published_raw", "")); c += 1
            put(ws, r, c, it.get("extract_status", "")); c += 1
            put(ws, r, c, best_url(it), link=True)
        ws.auto_filter.ref = f"A1:{gcl(len(cols))}{len(cands) + 1}"

    def add_top_tab(stream, cands, n=10):
        wl = stream == "watchlist"
        cols = [("Rank", 5), ("Score", 8), ("Breakdown", 30), ("Headline", 60)]
        if wl:
            cols += [("Company", 16)]
        cols += [("Keywords", 26), ("Corrob", 7), ("Source", 22), ("URL", 14)]
        ws = wb.create_sheet(f"{STREAM_LABEL[stream]} — top"[:31])
        header(ws, cols)
        for rank, it in enumerate(cands[:n], start=1):
            r, c = rank + 1, 1
            put(ws, r, c, rank); c += 1
            put(ws, r, c, it.get("score", 0)); c += 1
            put(ws, r, c, it.get("score_breakdown", ""), wrap=True); c += 1
            put(ws, r, c, it.get("title", ""), wrap=True); c += 1
            if wl:
                put(ws, r, c, it.get("matched_company", "")); c += 1
            put(ws, r, c, ", ".join(it.get("keywords_matched", []))); c += 1
            put(ws, r, c, it.get("cluster_size", 1)); c += 1
            put(ws, r, c, it.get("source", "")); c += 1
            put(ws, r, c, best_url(it), link=True)

    def add_by_company_tab(cands):
        groups: dict[str, list] = {}
        tier_of: dict[str, int] = {}
        for it in cands:
            co = it.get("matched_company") or ""
            if not co:
                continue
            groups.setdefault(co, []).append(it)
            tier_of[co] = it.get("company_tier", 9)
        if not groups:
            return
        ws = wb.create_sheet("Watchlist — by company")
        ws.cell(row=1, column=1, value="Watchlist coverage grouped by company (tier, then # stories)"
                ).font = Font(bold=True, size=11, color=NAVY)
        row = 3
        for co in sorted(groups, key=lambda k: (tier_of.get(k, 9), -len(groups[k]), k)):
            its = sorted(groups[co], key=lambda x: -x.get("score", 0))
            hc = ws.cell(row=row, column=1, value=f"{co}  (Tier {tier_of.get(co,'?')} · {len(its)} stories)")
            hc.font = Font(bold=True, color=WHITE, size=10)
            hc.fill = PatternFill("solid", fgColor=NAVY)
            row += 1
            for it in its:
                put(ws, row, 1, it.get("score", 0))
                put(ws, row, 2, it.get("title", ""), wrap=True)
                put(ws, row, 3, it.get("source", ""))
                put(ws, row, 4, best_url(it), link=True)
                row += 1
            row += 1
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 64
        ws.column_dimensions["C"].width = 22
        ws.column_dimensions["D"].width = 14

    for stream in ("ainews", "watchlist"):
        res = results.get(stream)
        if not res:
            continue
        cands = res.get("candidates", [])
        add_all_tab(stream, cands)
        add_top_tab(stream, cands)
        if stream == "watchlist":
            add_by_company_tab(cands)

    wb.save(path)
    return path


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="pboat RSS universe funnel")
    parser.add_argument("--stream", choices=["ainews", "watchlist", "both"],
                        default="both", help="Which stream(s) to generate")
    parser.add_argument("--hours", type=int, default=24,
                        help="Lookback window in hours (default 24)")
    parser.add_argument("--top", type=int, default=40,
                        help="Max candidates per stream (default 40)")
    parser.add_argument("--all-sources", action="store_true",
                        help="Keep off-allowlist publishers too (still tagged on_allowlist). "
                             "Default: drop them so the pool only contains trusted outlets.")
    parser.add_argument("--enrich-n", type=int, default=12,
                        help="How many top-scored candidates per stream to fetch full body "
                             "text for (default 12). 0 disables enrichment.")
    parser.add_argument("--no-enrich", action="store_true",
                        help="Skip full-text enrichment entirely (RSS titles+snippets only).")
    parser.add_argument("--no-playwright", action="store_true",
                        help="Don't use the Playwright fallback (requests-only extraction; "
                             "Google News redirect links will fall back to the RSS snippet).")
    args = parser.parse_args(argv)
    enrich_n = 0 if args.no_enrich else args.enrich_n
    use_playwright = not args.no_playwright

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    now_bkk = datetime.now(BKK)
    # Dashed YYYY-MM-DD — MUST match the engine's {TODAY} (Step 0.5 looks for
    # universe_{TODAY}_{stream}.json) and the article filenames (articles/{TODAY}-{stream}.md).
    date_str = now_bkk.strftime("%Y-%m-%d")

    # Load watchlist + the trusted-source allowlist (domains + per-section roles) once
    watchlist_keywords  = load_watchlist_keywords(WATCHLIST_PATH)
    watchlist_gnews     = load_watchlist_gnews_queries(WATCHLIST_PATH)
    watchlist_companies = load_watchlist_companies(WATCHLIST_PATH)
    trusted_domains     = load_trusted_domains(TRUSTED_SRC_PATH)
    source_roles        = load_source_roles(TRUSTED_SRC_PATH)
    trusted_only        = not args.all_sources
    print(f"[init] watchlist keywords loaded: {len(watchlist_keywords)}", flush=True)
    print(f"[init] watchlist GNews queries: {len(watchlist_gnews)}", flush=True)
    print(f"[init] watchlist companies: {len(watchlist_companies)}", flush=True)
    print(f"[init] trusted-source domains loaded: {len(trusted_domains)} "
          f"(filter {'ON' if trusted_only else 'OFF — --all-sources'}) · roles: {len(source_roles)}", flush=True)
    print(f"[init] enrichment: {'OFF' if enrich_n == 0 else f'top {enrich_n}/stream'} · "
          f"playwright {'available' if (use_playwright and PLAYWRIGHT_OK) else 'OFF'}", flush=True)

    streams = ["ainews", "watchlist"] if args.stream == "both" else [args.stream]

    results: dict[str, dict] = {}
    for stream in streams:
        result = run_stream(
            stream=stream,
            hours=args.hours,
            watchlist_keywords=watchlist_keywords,
            watchlist_gnews=watchlist_gnews,
            trusted_domains=trusted_domains,
            source_roles=source_roles,
            watchlist_companies=watchlist_companies,
            trusted_only=trusted_only,
            top_n=args.top,
            enrich_n=enrich_n,
            use_playwright=use_playwright,
        )
        results[stream] = result
        out_path = OUTPUT_DIR / f"universe_{date_str}_{stream}.json"
        out_path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"[{stream}] wrote → {out_path}", flush=True)

    # Combined ranked-universe workbook: a dated copy + a stable universe-latest.xlsx
    # (the brief footer links the stable one). Never let an Excel failure abort the run —
    # the JSON above is the load-bearing handoff; the Excel is a human-readable view of it.
    try:
        dated  = OUTPUT_DIR / f"universe_{date_str}.xlsx"
        latest = OUTPUT_DIR / "universe-latest.xlsx"
        export_universe_xlsx(results, dated)
        export_universe_xlsx(results, latest)
        print(f"[excel] wrote → {dated.name} + {latest.name}", flush=True)
    except Exception as e:
        print(f"[excel] WARN: could not write workbook ({type(e).__name__}: {e})", flush=True)

    print("\n[done]", flush=True)
    return 0


if __name__ == "__main__":
    sys.exit(main())
